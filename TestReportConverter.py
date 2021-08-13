'''Data Converter for AIBT'''
'''1. Search logs in folder.'''
'''2. Testlog Data Crawler.'''
'''3. .csv transfer to .xlsx & sheetnames'''


import os   # 模組 - 負責讀取資料夾跟檔案
import csv  # 模組 - 處理 CSV

from easygui import diropenbox  # 模組 - 直接導入選擇資料夾的UI畫面
from openpyxl import load_workbook  # 模組 - 處理 Excel

Question = input('是否執行轉檔程式? (Y/N) ')

if Question == 'Y': # 繼續執行
    path = diropenbox()  # 跳出對話窗 選擇 資料夾路徑

    #=================================================================================================
    #=================================== SerialNumber_Data Crawler ===================================
    #=================================================================================================
    def SNDataCrawler():
        files = os.listdir(path)  # 得到資料夾下的所有檔名稱
        for file in files:
            SerialNumber = file.split('.')
            #print(SerialNumber[0])  # 序號 -> 新增至sheetname

            # =================================================================================================
            # =================================== Open CSV to recode data ====================================
            # =================================================================================================
            with open(r'%s\%s'%(path, file), encoding='utf-8', newline='') as csvfile:
                Text = csv.reader(csvfile, delimiter=',', quotechar='|')
                Meas = []
                for row in Text:
                    Meas.append(row[13]) #指到 CSV的 Meas欄位
                #print(Meas[2]) #指到Meas的第三列

                # =================================================================================================
                # =================================== Open Excel to merge data ====================================
                # =================================================================================================
                WorkBook = load_workbook('TestReport.xlsx')
                filename = 'TestReport.xlsx'

                Sheet = WorkBook['Format']
                Tatget = WorkBook.copy_worksheet(Sheet)     # 記得要先建立有Forma格式的活頁簿
                Tatget.title = "%s"%(SerialNumber[0])


                # =================================================================================================
                # =================================== Filled in the table =========================================
                # =================================================================================================
                # 定值
                Tatget['D7'] = "No Short"
                Tatget['D8'] = "No Short"
                Tatget['D9'] = "No Short"

                Tatget['D15'] = "NA"
                Tatget['D19'] = "NA"
                Tatget['D24'] = "NA"
                Tatget['D26'] = "NA"
                Tatget['D28'] = "NA"
                Tatget['D30'] = "OK"
                Tatget['D31'] = "OK"
                Tatget['D32'] = "NA"

                # 待運算填值
                Tatget['A5'] = "SN: %s"%(SerialNumber[0])
                Tatget['D10'] = "%s"%(Meas[4])
                Tatget['D11'] = "%s" % (Meas[5])
                Tatget['D12'] = "%s" % (Meas[6])
                Tatget['D13'] = "%s" % (Meas[7])
                Tatget['D14'] = "%s" % (Meas[8])

                Tatget['D16'] = "%d" % (int(float(Meas[9])))    #將科學記號化繁為簡為整數, QC確認OK.
                Tatget['D17'] = "%s" % (Meas[10])
                Tatget['D18'] = "%s" % (Meas[11])
                Tatget['D20'] = "%s" % (Meas[12])
                Tatget['D21'] = "%s" % (Meas[13])
                Tatget['D22'] = "%s" % (Meas[14])
                Tatget['D23'] = "%s" % (Meas[15])
                Tatget['D25'] = "%s" % (Meas[16])
                Tatget['D27'] = "%s" % (Meas[17])
                Tatget['D29'] = "%s" % (Meas[18])
                Tatget['D33'] = "%s" % (Meas[21])
                Tatget['D34'] = "%s" % (Meas[22])
                Tatget['D35'] = "%s" % (Meas[23])


                WorkBook.save(filename = filename)
                print('%s is OK.'%(SerialNumber[0]))
                print(Meas[9], type(Meas[9]))
else:
    print('請先將檔案轉檔為Excel.')
    quit()


SNDataCrawler()
